# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'nac.ui'
#
# Created by: PyQt5 UI code generator 5.15.6
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt5 import QtCore, QtGui, QtWidgets
import os
from openpyxl import load_workbook


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(944, 715)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("icon/Petalm-logo.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setEnabled(True)
        self.tabWidget.setGeometry(QtCore.QRect(30, 30, 891, 611))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        self.tabWidget.setFont(font)
        self.tabWidget.setTabPosition(QtWidgets.QTabWidget.North)
        self.tabWidget.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.tabWidget.setElideMode(QtCore.Qt.ElideLeft)
        self.tabWidget.setTabBarAutoHide(True)
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.groupBox = QtWidgets.QGroupBox(self.tab)
        self.groupBox.setGeometry(QtCore.QRect(30, 30, 811, 511))
        self.groupBox.setObjectName("groupBox")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(40, 110, 71, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.title = QtWidgets.QLineEdit(self.groupBox)
        self.title.setGeometry(QtCore.QRect(230, 100, 481, 51))
        self.title.setText("")
        self.title.setClearButtonEnabled(True)
        self.title.setObjectName("title")
        self.filePath = QtWidgets.QLineEdit(self.groupBox)
        self.filePath.setGeometry(QtCore.QRect(230, 260, 481, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.filePath.setFont(font)
        self.filePath.setReadOnly(True)
        self.filePath.setClearButtonEnabled(False)
        self.filePath.setObjectName("filePath")
        self.findButton = QtWidgets.QPushButton(self.groupBox)
        self.findButton.setGeometry(QtCore.QRect(30, 260, 141, 51))
        self.findButton.setFlat(False)
        self.findButton.setObjectName("findButton")
        self.line = QtWidgets.QFrame(self.groupBox)
        self.line.setGeometry(QtCore.QRect(20, 340, 761, 20))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.genReport = QtWidgets.QPushButton(self.groupBox)
        self.genReport.setGeometry(QtCore.QRect(120, 410, 231, 61))
        self.genReport.setObjectName("genReport")
        self.openReports = QtWidgets.QPushButton(self.groupBox)
        self.openReports.setGeometry(QtCore.QRect(440, 410, 231, 61))
        self.openReports.setObjectName("openReports")
        self.line_2 = QtWidgets.QFrame(self.groupBox)
        self.line_2.setGeometry(QtCore.QRect(30, 50, 761, 20))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.tabWidget.addTab(self.tab, "")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 944, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Petalm Africa Services"))
        self.groupBox.setTitle(_translate("MainWindow", "Enter The Following"))
        self.label.setText(_translate("MainWindow", "Title:"))
        self.filePath.setText(_translate("MainWindow", "Select File.."))
        self.findButton.setText(_translate("MainWindow", "Find:"))
        self.genReport.setText(_translate("MainWindow", "Genarate Report"))
        self.openReports.setText(_translate("MainWindow", "Open Reports"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", "Nac Reports"))

#----------- Functionality Mod Start-----------------#
        self.genReport.clicked.connect(self.get_data)
        self.findButton.clicked.connect(self.findFunction)
        self.openReports.clicked.connect(self.openFunction)

    def openFunction(self):
            path = os.path.abspath('Nac Reports')
            document = QtWidgets.QFileDialog.getOpenFileName(self.tabWidget,'Open An F16 Report:',path, filter = 'f16 Files (*.f16 )')
           # print(document)
    
    def findFunction(self):
            path = os.path.abspath('documents')
            document = QtWidgets.QFileDialog.getOpenFileName(self.tabWidget,'Select An Excel File:',path, filter = 'Excel Files (*.xlsx *.xlsm *.xltx *.xltm)')
           # print(document)
            self.filePath.setText(document[0])
            self.paths =document[0] 
        
    def process(self, wording, spacing):
        max_len = len(wording)
        if max_len > spacing:
            return wording[0:spacing]
        padding = spacing-max_len
        for i in range(0, padding):
            wording = wording+" "
        return str(wording)

    def fix(self, wording):
        spacing = 12
        max_len = len(wording)
        if max_len > spacing:
            return wording[0:spacing]
        padding = spacing-max_len
        for i in range(0, padding):
            wording = wording+"*"
        return str(wording)

    def get_data(self):
        try:
            print('title: ',self.title.text())
            title = self.title.text()
            path = self.paths
            wb = load_workbook(path)
            ws = wb.active
            cols = [x for x in range(2, int(ws.max_column)-2)]
            
            with open('Nac Reports/'+title+'.f16','a', encoding = 'utf-8') as f:
                f.write('HDR20211231                    000200011098NATIONAL AIDS COUNCIL                   202201260112120100                                                                                                                                                                                                                                                                                                                                                                ')
                for x in (range(2,ws.max_row)):
                    
                    f.write('\n')
                    f.write(self.process('TAX',9))
                    for i in cols:
                        
                        if i <5 and i!=4:
                            f.writelines(self.process(str(ws.cell(row=x, column=i).value),14))
                        elif i ==4:
                            f.writelines(self.process(str(ws.cell(row=x, column=i).value),25))
                        else:
                            valueA = str(ws.cell(row=x, column=i).value).replace('.', '')
                            value = valueA.replace(' ', '')
                            #count = count+str(value).replace('-', '') 
                            f.writelines(self.fix(str(value).replace('-', '')))
                f.write('\n')
                f.write('TRL000394000044353818552000012727355693')
                self.filePath.setText('Select File..')
                self.title.setText('')
                return QtWidgets.QMessageBox.about(self.tabWidget,'Success','Report Generated Successfully')
        except Exception as e:
            print(e)
            return QtWidgets.QMessageBox.about(self.tabWidget,'Error','Report Generated Failed, Please Select an Excel File')


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('fusion')
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
